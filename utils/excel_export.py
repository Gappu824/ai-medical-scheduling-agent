"""
Enhanced Excel Export with Complete Reminder System Data & Visual Highlighting
RagaAI Assignment - Export all appointment data with visual cues for status
"""

import pandas as pd
import logging
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class EnhancedExcelExporter:
    """Enhanced Excel exporter that includes visual highlighting for status"""
    
    def __init__(self):
        self.exports_dir = Path("exports")
        self.exports_dir.mkdir(exist_ok=True)
        self.db_path = "medical_scheduling.db"
        
        # Define styles for highlighting
        self.confirmed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.cancelled_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.cancelled_font = Font(color="9C0006", strike=True)
        
        logger.info("Enhanced Excel Exporter with visual highlighting initialized")
    
    def export_complete_appointment_data(self) -> str:
        """
        Export complete appointment data including visual highlights for status.
        """
        try:
            conn = sqlite3.connect(self.db_path)
            
            main_query = """
            SELECT 
                a.id as appointment_id,
                p.first_name || ' ' || p.last_name as patient_name,
                a.appointment_datetime,
                a.doctor,
                a.location,
                a.duration,
                a.status,
                p.patient_type,
                p.email,
                p.phone,
                p.insurance_carrier
            FROM appointments a
            JOIN patients p ON a.patient_id = p.id
            ORDER BY a.appointment_datetime DESC
            """
            
            appointments_df = pd.read_sql_query(main_query, conn)
            conn.close()
            
            if appointments_df.empty:
                logger.warning("No appointments found to export.")
                return None
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"appointments_report_{timestamp}.xlsx"
            filepath = self.exports_dir / filename
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Appointments"
            
            # Append DataFrame to worksheet
            for r in dataframe_to_rows(appointments_df, index=False, header=True):
                ws.append(r)
            
            # Apply conditional formatting
            header_row = ws[1]
            status_col_index = -1
            for idx, cell in enumerate(header_row):
                if cell.value == 'status':
                    status_col_index = idx + 1
                    break

            if status_col_index != -1:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=status_col_index, max_col=status_col_index):
                    for cell in row:
                        if cell.value.lower() == 'confirmed':
                            for r_cell in ws[cell.row]:
                                r_cell.fill = self.confirmed_fill
                        elif cell.value.lower() == 'cancelled':
                            for r_cell in ws[cell.row]:
                                r_cell.fill = self.cancelled_fill
                                r_cell.font = self.cancelled_font

            wb.save(filepath)
            
            logger.info(f"✅ Complete appointment export with visual highlighting created: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ Complete appointment export failed: {e}", exc_info=True)
            return None

# For backward compatibility
ExcelExporter = EnhancedExcelExporter